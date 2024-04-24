import streamlit as st
import cv2
import pytesseract
from openpyxl import Workbook, load_workbook
import base64
import numpy as np
import os

# Function to process images and extract information
def process_image(image):
    # Convert image to grayscale
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)

    # Perform OCR on the grayscale image
    data = pytesseract.image_to_string(gray)

    # Extract relevant information from OCR results
    patient_name = None
    contact = None
    address = None
    date = None

    data = data.split("\n")
    for item in data:
        if 'Patient Name:' in item:
            patient_name = item.split(':')[-1].strip()
        elif 'Contact:' in item:
            contact = item.split(':')[-1].strip()
        elif 'Address:' in item:
            address = item.split(':')[-1].strip()
        elif 'Date:' in item:
            date = item.split(':')[1].strip()

    return patient_name, contact, address, date

# Function to append extracted information to an Excel file
def append_to_excel(extracted_info):
    try:
        excel_path = "output/extracted_data.xlsx"
        if os.path.exists(excel_path):  # Check if Excel file already exists
            wb = load_workbook(excel_path)
            ws = wb.active
        else:  # If Excel file doesn't exist, create a new one
            wb = Workbook()
            ws = wb.active
            ws.append(["Patient Name", "Contact", "Address", "Date"])  # Add headers if creating a new file

        # Append the extracted information to the worksheet
        ws.append([extracted_info['Patient Name'], extracted_info['Contact'], extracted_info['Address'], extracted_info['Date']])

        # Save the workbook
        wb.save(excel_path)

        return excel_path
    except Exception as e:
        st.error(f"Error while saving data to Excel: {e}")
        return None

# Streamlit app
def main():
    st.title('Image Processing and Data Extraction')

    # File upload widget
    uploaded_file = st.file_uploader("Upload an image", type=["jpg", "png", "jpeg"])

    if uploaded_file is not None:
        # Read the uploaded image
        file_bytes = uploaded_file.getvalue()
        nparr = np.frombuffer(file_bytes, np.uint8)
        image = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

        # Display the uploaded image
        st.image(image, caption='Uploaded Image', use_column_width=True)

        # Process the uploaded image
        patient_name, contact, address, date = process_image(image)

        # Display the extracted information
        st.write("Extracted Information:")
        st.write("Patient Name:", patient_name)
        st.write("Contact:", contact)
        st.write("Address:", address)
        st.write("Date:", date)

        # Save the extracted information to an Excel file
        extracted_info = {'Patient Name': patient_name, 'Contact': contact, 'Address': address, 'Date': date}
        excel_path = append_to_excel(extracted_info)

        # Provide a download link for the Excel file
        if excel_path:
            st.write('Download the extracted data [here](data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{})'.format(
                base64.b64encode(open(excel_path, "rb").read()).decode()
            ), unsafe_allow_html=True)

if __name__ == '__main__':
    main()
